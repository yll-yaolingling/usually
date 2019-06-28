import openpyxl
from openpyxl import load_workbook

wb=load_workbook("D:/1.xlsx")
ws=wb["Sheet1"]

for i in range(12,29):
    ws.cell(i,2).value="u"
if ws.cell(22,2).value=="u":
    ws.cell(22, 2).value="a"
wb.save("D:/1.xlsx")
wb.close()

